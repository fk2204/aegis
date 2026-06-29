"""A/R aging extractor — Excel / CSV / PDF.

Three paths share a single ``ARAgingResult`` shape:

* ``extract_ar_aging_excel`` — openpyxl, deterministic header match on
  the first sheet. Headers we accept (case-insensitive substring):
  current / 0-30 / 31-60 / 30-60 / 61-90 / 60-90 / 91+ / 90+ / 120+.
* ``extract_ar_aging_csv`` — same header logic on csv.DictReader rows.
* ``extract_ar_aging_pdf`` — Bedrock forced-tool-use call (vision +
  text). Used only when the operator uploads a printed PDF aging
  report; deterministic paths preferred when the operator can export
  the underlying xlsx / csv.

Money is ``Decimal`` everywhere. ``concentration_pct`` is the share of
``total_outstanding`` held by the single largest debtor — surfaced on
the dossier as a soft concern at >40%.
"""

from __future__ import annotations

import base64
import csv
import re
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Final, Protocol

import openpyxl

from aegis.logger import get_logger

_log = get_logger(__name__)


# ----------------------------------------------------------------------
# Detection
# ----------------------------------------------------------------------


# Filename substring → A/R aging. Case-insensitive. ``invoice`` is
# DELIBERATELY omitted — that token belongs to the equipment-invoice
# parser's detection (parser/equipment/) and would collide here.
_AR_FILENAME_TOKENS: Final[tuple[str, ...]] = (
    "aging",
    "receivable",
    "ar_",
    "a_r",
    "a/r ",
)


def detect_ar_aging_filename(filename: str) -> bool:
    """Return True when the filename looks like an A/R aging report.

    Pure substring match on the lowercased basename. Callers pass the
    BASENAME (not the full path) to avoid the path itself producing a
    false positive.
    """
    if not filename:
        return False
    name = filename.lower()
    return any(token in name for token in _AR_FILENAME_TOKENS)


# ----------------------------------------------------------------------
# Result shape
# ----------------------------------------------------------------------


@dataclass(frozen=True)
class ARAgingResult:
    """Validated A/R aging extraction. Money in Decimal."""

    total_outstanding: Decimal
    current_amount: Decimal
    days_30_60: Decimal
    days_60_90: Decimal
    days_90_plus: Decimal
    top_debtors: list[dict[str, Any]] = field(default_factory=list)
    debtor_count: int = 0
    concentration_pct: Decimal = Decimal("0.00")
    extracted_at: datetime = field(default_factory=lambda: datetime.now(UTC))
    error: str | None = None


# ----------------------------------------------------------------------
# Header matching — Excel + CSV share this
# ----------------------------------------------------------------------


# Header substring → ARAgingResult bucket name. Order matters: the
# longest / most-specific match wins.
_BUCKET_HEADER_PATTERNS: Final[tuple[tuple[str, str], ...]] = (
    ("120+", "days_90_plus"),  # fold 120+ into 90+
    ("91+", "days_90_plus"),
    ("90+", "days_90_plus"),
    ("over 90", "days_90_plus"),
    ("91 and over", "days_90_plus"),
    ("61-90", "days_60_90"),
    ("60-90", "days_60_90"),
    ("31-60", "days_30_60"),
    ("30-60", "days_30_60"),
    ("0-30", "current_amount"),
    ("1-30", "current_amount"),
    ("current", "current_amount"),
    ("not yet due", "current_amount"),
)


# Header substring → debtor-name column. We accept several common labels.
_DEBTOR_HEADER_TOKENS: Final[tuple[str, ...]] = (
    "customer",
    "debtor",
    "name",
    "account",
    "client",
)


# Header substring → total column. Optional — when present, we use it
# as the row's total instead of summing the buckets (handles reports
# where the buckets don't account for the entire balance, e.g. credit
# memos).
_TOTAL_HEADER_TOKENS: Final[tuple[str, ...]] = (
    "total",
    "balance",
    "outstanding",
)


def _classify_header(header: str) -> str | None:
    """Return the bucket name for an aging-column header, or None.

    Bucket names match ``ARAgingResult`` field names so the caller can
    ``setattr`` directly.
    """
    h = header.lower().strip()
    for token, bucket in _BUCKET_HEADER_PATTERNS:
        if token in h:
            return bucket
    return None


def _is_debtor_header(header: str) -> bool:
    h = header.lower().strip()
    return any(token in h for token in _DEBTOR_HEADER_TOKENS)


def _is_total_header(header: str) -> bool:
    h = header.lower().strip()
    if any(b in h for b in ("30", "60", "90", "120", "current")):
        # Buckets often contain "total" in a wider phrase ("31-60 days
        # totals"); they're not THE row-total column.
        return False
    return any(token in h for token in _TOTAL_HEADER_TOKENS)


# ----------------------------------------------------------------------
# Money parsing
# ----------------------------------------------------------------------


_MONEY_STRIP_RE: Final[re.Pattern[str]] = re.compile(r"[\s,$]")


def _parse_money(raw: Any) -> Decimal:  # noqa: ANN401 — heterogeneous Excel/CSV cell type
    """Coerce an Excel / CSV cell to ``Decimal``. Empty / non-numeric
    -> ``Decimal("0")``. Negative values preserved (credit memos).
    Parens-negative ("(1,234.56)") supported.
    """
    if raw is None:
        return Decimal("0")
    if isinstance(raw, (int, float, Decimal)):
        try:
            return Decimal(str(raw))
        except (InvalidOperation, ValueError):
            return Decimal("0")
    s = str(raw).strip()
    if not s:
        return Decimal("0")
    negative = s.startswith("(") and s.endswith(")")
    if negative:
        s = s[1:-1]
    s = _MONEY_STRIP_RE.sub("", s)
    if not s:
        return Decimal("0")
    try:
        value = Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")
    return -value if negative else value


# ----------------------------------------------------------------------
# Excel extraction
# ----------------------------------------------------------------------


def _build_result(
    rows: list[dict[str, Any]],
    *,
    debtor_key: str | None,
    total_key: str | None,
) -> ARAgingResult:
    """Pure aggregation: rows → ARAgingResult. Shared between Excel +
    CSV paths so they produce byte-identical output for the same data.
    """
    bucket_keys = {"current_amount", "days_30_60", "days_60_90", "days_90_plus"}
    sums: dict[str, Decimal] = {k: Decimal("0") for k in bucket_keys}
    debtors: list[dict[str, Any]] = []
    for r in rows:
        row_total = Decimal("0")
        for k in bucket_keys:
            v = _parse_money(r.get(k))
            sums[k] += v
            row_total += v
        if total_key is not None:
            override = _parse_money(r.get(total_key))
            if override > 0:
                row_total = override
        if debtor_key and row_total != 0:
            name = str(r.get(debtor_key) or "").strip()
            if name:
                debtors.append({"name": name, "amount": row_total})

    if total_key is not None:
        total = sum((_parse_money(r.get(total_key)) for r in rows), start=Decimal("0"))
        if total <= 0:
            total = sum(sums.values(), start=Decimal("0"))
    else:
        total = sum(sums.values(), start=Decimal("0"))

    debtors.sort(key=lambda d: d["amount"], reverse=True)
    top = debtors[:5]
    if total > 0:
        for d in top:
            d["pct_of_total"] = (d["amount"] / total * Decimal("100")).quantize(Decimal("0.01"))
    concentration = (
        (top[0]["amount"] / total * Decimal("100")).quantize(Decimal("0.01"))
        if top and total > 0
        else Decimal("0.00")
    )

    return ARAgingResult(
        total_outstanding=total,
        current_amount=sums["current_amount"],
        days_30_60=sums["days_30_60"],
        days_60_90=sums["days_60_90"],
        days_90_plus=sums["days_90_plus"],
        top_debtors=[{**d, "amount": str(d["amount"])} for d in top],
        debtor_count=len(debtors),
        concentration_pct=concentration,
    )


def extract_ar_aging_excel(file_path: str | Path) -> ARAgingResult:
    """Parse an .xlsx A/R aging file. First sheet, header row 1."""
    path = Path(file_path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return ARAgingResult(
                total_outstanding=Decimal("0"),
                current_amount=Decimal("0"),
                days_30_60=Decimal("0"),
                days_60_90=Decimal("0"),
                days_90_plus=Decimal("0"),
                error="empty workbook",
            )
        iter_rows = ws.iter_rows(values_only=True)
        try:
            headers_row = next(iter_rows)
        except StopIteration:
            return ARAgingResult(
                total_outstanding=Decimal("0"),
                current_amount=Decimal("0"),
                days_30_60=Decimal("0"),
                days_60_90=Decimal("0"),
                days_90_plus=Decimal("0"),
                error="no header row",
            )

        bucket_columns: dict[int, str] = {}
        debtor_col: int | None = None
        total_col: int | None = None
        for idx, raw_h in enumerate(headers_row):
            if raw_h is None:
                continue
            h = str(raw_h)
            bucket = _classify_header(h)
            if bucket is not None:
                bucket_columns[idx] = bucket
                continue
            if debtor_col is None and _is_debtor_header(h):
                debtor_col = idx
                continue
            if total_col is None and _is_total_header(h):
                total_col = idx

        if not bucket_columns:
            return ARAgingResult(
                total_outstanding=Decimal("0"),
                current_amount=Decimal("0"),
                days_30_60=Decimal("0"),
                days_60_90=Decimal("0"),
                days_90_plus=Decimal("0"),
                error="no aging-bucket columns matched in header",
            )

        rows: list[dict[str, Any]] = []
        for row in iter_rows:
            mapped: dict[str, Any] = {}
            for idx, bucket in bucket_columns.items():
                if idx < len(row):
                    mapped[bucket] = row[idx]
            if debtor_col is not None and debtor_col < len(row):
                mapped["__debtor__"] = row[debtor_col]
            if total_col is not None and total_col < len(row):
                mapped["__total__"] = row[total_col]
            rows.append(mapped)

        return _build_result(
            rows,
            debtor_key="__debtor__" if debtor_col is not None else None,
            total_key="__total__" if total_col is not None else None,
        )
    finally:
        wb.close()


# ----------------------------------------------------------------------
# CSV extraction
# ----------------------------------------------------------------------


def extract_ar_aging_csv(file_path: str | Path) -> ARAgingResult:
    """Parse a CSV A/R aging file using header-row detection."""
    path = Path(file_path)
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            headers = next(reader)
        except StopIteration:
            return ARAgingResult(
                total_outstanding=Decimal("0"),
                current_amount=Decimal("0"),
                days_30_60=Decimal("0"),
                days_60_90=Decimal("0"),
                days_90_plus=Decimal("0"),
                error="empty CSV",
            )

        bucket_columns: dict[int, str] = {}
        debtor_col: int | None = None
        total_col: int | None = None
        for idx, raw_h in enumerate(headers):
            bucket = _classify_header(raw_h)
            if bucket is not None:
                bucket_columns[idx] = bucket
                continue
            if debtor_col is None and _is_debtor_header(raw_h):
                debtor_col = idx
                continue
            if total_col is None and _is_total_header(raw_h):
                total_col = idx

        if not bucket_columns:
            return ARAgingResult(
                total_outstanding=Decimal("0"),
                current_amount=Decimal("0"),
                days_30_60=Decimal("0"),
                days_60_90=Decimal("0"),
                days_90_plus=Decimal("0"),
                error="no aging-bucket columns matched in header",
            )

        rows: list[dict[str, Any]] = []
        for raw_row in reader:
            mapped: dict[str, Any] = {}
            for idx, bucket in bucket_columns.items():
                if idx < len(raw_row):
                    mapped[bucket] = raw_row[idx]
            if debtor_col is not None and debtor_col < len(raw_row):
                mapped["__debtor__"] = raw_row[debtor_col]
            if total_col is not None and total_col < len(raw_row):
                mapped["__total__"] = raw_row[total_col]
            rows.append(mapped)

        return _build_result(
            rows,
            debtor_key="__debtor__" if debtor_col is not None else None,
            total_key="__total__" if total_col is not None else None,
        )


# ----------------------------------------------------------------------
# PDF extraction — Bedrock vision
# ----------------------------------------------------------------------


class _LLMClient(Protocol):
    def invoke_tool_json(
        self,
        *,
        system_prompt: str,
        user_prompt: str,
        tool_name: str,
        tool_schema: dict[str, Any],
        max_tokens: int,
        temperature: float,
        pdf_b64: str | None = None,
    ) -> tuple[dict[str, Any], str]: ...


_PDF_TOOL_SCHEMA: Final[dict[str, Any]] = {
    "type": "object",
    "properties": {
        "total_outstanding": {"type": "string"},
        "current_amount": {"type": "string"},
        "days_30_60": {"type": "string"},
        "days_60_90": {"type": "string"},
        "days_90_plus": {"type": "string"},
        "top_debtors": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "amount": {"type": "string"},
                },
                "required": ["name", "amount"],
                "additionalProperties": False,
            },
            "maxItems": 10,
        },
        "debtor_count": {"type": "integer"},
    },
    "required": [
        "total_outstanding",
        "current_amount",
        "days_30_60",
        "days_60_90",
        "days_90_plus",
    ],
    "additionalProperties": False,
}


_PDF_SYSTEM_PROMPT: Final[str] = (
    "Extract A/R aging totals from this report. Return money figures as "
    "Decimal-safe strings (no $, no comma) — e.g. '125000.00', not "
    "'$125,000'. Sum any 120+ bucket into days_90_plus. List the top "
    "five debtors by outstanding balance with their dollar amount. NEVER "
    "invent figures — if a bucket is not shown, emit '0' for it; if the "
    "report is illegible, return an error string. Be deterministic."
)


def extract_ar_aging_pdf(
    file_path: str | Path,
    *,
    llm_client: _LLMClient,
) -> ARAgingResult:
    """Parse a PDF A/R aging report via Bedrock forced tool-use."""
    path = Path(file_path)
    pdf_bytes = path.read_bytes()
    pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("ascii")
    raw, _model_id = llm_client.invoke_tool_json(
        system_prompt=_PDF_SYSTEM_PROMPT,
        user_prompt="Extract the A/R aging totals and top debtors.",
        tool_name="record_ar_aging",
        tool_schema=_PDF_TOOL_SCHEMA,
        max_tokens=1024,
        temperature=0.0,
        pdf_b64=pdf_b64,
    )

    total = _parse_money(raw.get("total_outstanding"))
    current = _parse_money(raw.get("current_amount"))
    d30 = _parse_money(raw.get("days_30_60"))
    d60 = _parse_money(raw.get("days_60_90"))
    d90 = _parse_money(raw.get("days_90_plus"))

    if total <= 0:
        total = current + d30 + d60 + d90

    raw_debtors = raw.get("top_debtors") or []
    top: list[dict[str, Any]] = []
    for d in raw_debtors:
        if not isinstance(d, dict):
            continue
        name = str(d.get("name") or "").strip()
        amount = _parse_money(d.get("amount"))
        if not name or amount <= 0:
            continue
        entry: dict[str, Any] = {"name": name, "amount": str(amount)}
        if total > 0:
            entry["pct_of_total"] = (amount / total * Decimal("100")).quantize(Decimal("0.01"))
        top.append(entry)
    top.sort(key=lambda d: _parse_money(d["amount"]), reverse=True)
    top = top[:5]
    concentration = (
        (_parse_money(top[0]["amount"]) / total * Decimal("100")).quantize(Decimal("0.01"))
        if top and total > 0
        else Decimal("0.00")
    )

    raw_count = raw.get("debtor_count")
    debtor_count = int(raw_count) if isinstance(raw_count, int) else len(top)

    return ARAgingResult(
        total_outstanding=total,
        current_amount=current,
        days_30_60=d30,
        days_60_90=d60,
        days_90_plus=d90,
        top_debtors=top,
        debtor_count=debtor_count,
        concentration_pct=concentration,
    )
