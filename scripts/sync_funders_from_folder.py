"""Sync funder guidelines from a local folder into AEGIS.

Expected folder layout (default = the operator's Windows OneDrive path,
configured via ``settings.funders_folder_path``)::

    Funders/                          ← --folder / FUNDERS_FOLDER_PATH
      Highland Hill Capital/
        guidelines.pdf
        rate_card.xlsx
        screenshot.png
      Logic Advance/
        underwriting_matrix.docx
      New Funder/                     ← auto-detected → INSERT
        criteria.png

Per immediate subfolder the script processes **every** supported file
(``.pdf .docx .doc .xlsx .xls .jpg .jpeg .png .txt``), extracts each
through the canonical sanitiser + confidence-floor path, and merges the
per-file results into one extraction (first non-None wins for scalars,
dedup-union for list fields). A SHA-256 of every supported file's
bytes (deterministic name-sorted order) is stored at
``funders.guidelines_data._file_hash``; re-running with an unchanged
folder is a no-op. ANY file added, removed, or modified re-triggers
extraction for that funder.

Merge priority for first-non-None wins: PDFs are processed first, then
text-extracted formats (DOCX/XLSX/TXT), then images. Within each tier,
largest file first. Rationale: PDFs are the most-authoritative funder
guidelines source; the operator's screenshots are last-resort
supplements. The merge log is persisted at
``guidelines_data._files = [{name, action, …}]`` so the operator can
audit which file contributed which fields.

Bedrock routing per file type:

* PDF + images (PNG / JPG / JPEG) — the canonical AEGIS extractor
  ``aegis.funders.guidelines_extract.extract_guidelines_from_pdf``
  (and its sibling vision path for images). That extractor owns the
  sanitiser + per-field 0.5 confidence floor; we route through it so
  the script can't drift from the production policy.
* DOCX (mammoth → plaintext), XLSX (openpyxl → CSV-ish dump), TXT
  (utf-8) — a thin direct call to
  ``BedrockClient.classify_batch_json`` using the same prompt the
  canonical extractor uses. The same sanitiser is applied AEGIS-side
  before persistence so the confidence floor still holds.

Images above ``_IMAGE_MAX_BYTES`` (5 MB — Bedrock vision per-image
limit) are skipped with a logged WARN; the merge proceeds on the
remaining files. PDF + DOCX + XLSX + TXT use the canonical extractor's
own 25 MB ceiling.

Operator-confirmation discipline — CLAUDE.md "Extraction & automation
assists, never replaces judgment" — normally bans auto-writes to the
live FunderRow columns. This script is the explicit operator-approved
exception: the folder lives on the operator's own machine, contents are
their own funder criteria sheets (not Bedrock guessing from a stranger's
PDF), so per-funder INSERT seeds the live columns
(``min_revenue``, ``min_fico``, ``min_tib_months``, ``max_positions``,
``allows_stacking``, ``deal_types_accepted``, ``excluded_states``,
``excluded_industries``) directly. The structured JSONB
``guidelines_data`` blob remains the source of truth for the operator
to review; the live columns are a convenience pre-fill the operator can
still override via the funder edit UI.

Exit codes:

* ``0`` — clean run.
* ``1`` — config error (folder missing, supabase init failed).
* ``3`` — one or more per-funder errors (Bedrock failure, malformed
  file, Supabase write failure). Other funders still processed.

Usage::

    python scripts/sync_funders_from_folder.py                   # dry-run
    python scripts/sync_funders_from_folder.py --apply           # writes
    python scripts/sync_funders_from_folder.py --folder PATH     # override
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import os
import sys
import traceback
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast

# Make the script runnable directly from the repo root without a
# package install (matches sibling scripts).
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT / "src") not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT / "src"))


EXIT_OK = 0
EXIT_CONFIG = 1
EXIT_PER_FUNDER_ERRORS = 3


_SUPPORTED_EXTS: frozenset[str] = frozenset(
    {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".jpg", ".jpeg", ".png", ".txt"}
)
_PDF_EXTS: frozenset[str] = frozenset({".pdf"})
_IMAGE_EXTS: frozenset[str] = frozenset({".jpg", ".jpeg", ".png"})
_DOCX_EXTS: frozenset[str] = frozenset({".docx", ".doc"})
_XLSX_EXTS: frozenset[str] = frozenset({".xlsx", ".xls"})
_TXT_EXTS: frozenset[str] = frozenset({".txt"})

# Extension priority for the multi-file merge: PDFs (most-authoritative
# guidelines source) first, then text-extracted office formats, then
# images (often partial screenshots). Lower number = higher priority;
# first-non-None merge means PDFs win on field collisions.
_EXTENSION_PRIORITY: dict[str, int] = {
    ".pdf": 0,
    ".docx": 1,
    ".doc": 1,
    ".xlsx": 1,
    ".xls": 1,
    ".txt": 1,
    ".png": 2,
    ".jpg": 2,
    ".jpeg": 2,
}

# Bedrock vision accepts up to 5 MB per image (model card). PDFs go
# through the document content block which has its own 25 MB ceiling
# enforced inside ``extract_guidelines_from_pdf``.
_IMAGE_MAX_BYTES: int = 5 * 1024 * 1024

# Pydantic field names that take dedup-union semantics during merge.
# Anything not in this set takes first-non-None (priority order from
# ``_get_all_supported_files``). ``notes`` is treated specially —
# first-non-empty wins (empty string is the model default and means
# "no notes from this file").
_MERGE_LIST_FIELDS: frozenset[str] = frozenset({"excluded_industries", "excluded_states"})


def _load_dotenv() -> None:
    """Load ``.env`` from the repo root without overwriting existing env.

    Mirrors ``scripts/audit_funders_table.py``. The Settings class also
    auto-loads ``.env`` via pydantic-settings but does so RELATIVE to the
    process CWD, which fails when the script is invoked from a deeper
    directory. Explicit load keeps behavior stable.
    """
    for path in (_REPO_ROOT / ".env", _REPO_ROOT / ".env.local"):
        if not path.exists():
            continue
        for raw in path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__.split("\n", 1)[0])
    parser.add_argument(
        "--folder",
        type=str,
        default=None,
        help=(
            "Override the funders folder path. Defaults to "
            "settings.funders_folder_path (see src/aegis/config.py)."
        ),
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--apply",
        action="store_true",
        help="Actually write to Supabase. Default is dry-run.",
    )
    mode.add_argument(
        "--dry-run",
        action="store_true",
        help="Classify each subfolder; no DB writes. Default.",
    )
    return parser.parse_args()


def _sha256_file(path: Path) -> str:
    """SHA-256 hex digest of a file's bytes, read in 64 KiB chunks."""
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(64 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _get_all_supported_files(folder: Path) -> list[Path]:
    """Return every supported file in ``folder``, in merge-priority order.

    Immediate children only — sub-subfolders are ignored. Files are
    sorted by extension priority first (PDF > text-extracted >
    images), then by size descending within each tier so the heaviest
    PDF (most likely to be the canonical guidelines doc) gets the
    first-non-None slot. Word lock files (``~$…``) and hidden files
    (``.…``) are excluded.
    """
    files: list[Path] = []
    for entry in folder.iterdir():
        if not entry.is_file():
            continue
        if entry.name.startswith("~$") or entry.name.startswith("."):
            continue
        if entry.suffix.lower() in _SUPPORTED_EXTS:
            files.append(entry)
    files.sort(
        key=lambda p: (
            _EXTENSION_PRIORITY.get(p.suffix.lower(), 99),
            -p.stat().st_size,
        )
    )
    return files


def _folder_hash(folder: Path) -> str:
    """SHA-256 over every supported file in ``folder``.

    Deterministic — files are visited in name-sorted order (not the
    merge-priority order from ``_get_all_supported_files`` — name-sort
    is stable across file-system iteration quirks). Each file's name
    is mixed in alongside its bytes so a rename also bumps the hash.

    Changes when ANY file in the folder changes, so the
    ``stored_hash == folder_hash`` short-circuit re-triggers extraction
    whenever the operator adds, removes, renames, or modifies any
    contributing file.
    """
    digest = hashlib.sha256()
    files = sorted(_get_all_supported_files(folder), key=lambda p: p.name.lower())
    for f in files:
        digest.update(f.name.encode("utf-8"))
        with f.open("rb") as fh:
            for chunk in iter(lambda fh=fh: fh.read(64 * 1024), b""):  # type: ignore[misc]
                digest.update(chunk)
    return digest.hexdigest()


def _format_file_size(size_bytes: int) -> str:
    """Human-readable size for dry-run output (96K, 2.9M, 1.1G)."""
    if size_bytes < 1024:
        return f"{size_bytes}B"
    if size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.0f}K"
    if size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f}M"
    return f"{size_bytes / (1024 * 1024 * 1024):.1f}G"


def _content_label_for(path: Path) -> str:
    """Short label for dry-run output ("PDF", "DOCX", "XLSX", "TXT", "PNG", "JPG")."""
    ext = path.suffix.lower()
    if ext in _PDF_EXTS:
        return "PDF"
    if ext in _DOCX_EXTS:
        return ext.lstrip(".").upper()
    if ext in _XLSX_EXTS:
        return ext.lstrip(".").upper()
    if ext in _TXT_EXTS:
        return "TXT"
    if ext in _IMAGE_EXTS:
        return ext.lstrip(".").upper()
    return ext.lstrip(".") or "?"


def _xlsx_to_text(path: Path) -> str:
    """Dump every sheet of an xlsx file as CSV-ish plaintext.

    openpyxl is already an AEGIS dependency (A/R aging parser). Read-only
    mode keeps memory bounded on large workbooks. Empty / formula-only
    cells become empty strings, matching the existing A/R extractor.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    buf = io.StringIO()
    for sheet_name in wb.sheetnames:
        buf.write(f"# SHEET: {sheet_name}\n")
        ws = wb[sheet_name]
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if cell is None else str(cell) for cell in row])
        buf.write("\n")
    return buf.getvalue()


def _docx_to_text(path: Path) -> str:
    """Convert a docx file to plaintext via mammoth.

    Mammoth strips Word styling (which is what we want — Bedrock works
    better on clean text). ``.doc`` (binary legacy format) is NOT
    supported by mammoth; if the operator drops a ``.doc`` in we log it
    and skip rather than half-process.
    """
    import mammoth

    with path.open("rb") as fh:
        result = mammoth.extract_raw_text(fh)
    raw_value: object = result.value
    return raw_value if isinstance(raw_value, str) else str(raw_value)


def _extract_from_text(
    text: str,
    llm: Any,  # noqa: ANN401 — BedrockClient duck-typed for test stubs
) -> Any:  # noqa: ANN401 — returns aegis.funders.guidelines_extract.FunderGuidelinesExtraction
    """Extract guidelines from plain text via the canonical sanitiser.

    Reuses ``GUIDELINES_EXTRACTION_PROMPT`` + ``_sanitise_extraction`` +
    ``FunderGuidelinesExtraction`` from
    ``aegis.funders.guidelines_extract`` so the confidence floor and the
    coercion rules don't drift. The only thing this function does that
    the canonical PDF entry point doesn't is route text through
    ``classify_batch_json`` instead of the document-block call (Bedrock
    has no native "plain text" envelope for the document content type).
    """
    from pydantic import ValidationError

    from aegis.funders.guidelines_extract import (
        GUIDELINES_EXTRACTION_PROMPT,
        FunderGuidelinesExtraction,
        GuidelinesExtractionError,
        _sanitise_extraction,
    )

    full_prompt = f"{GUIDELINES_EXTRACTION_PROMPT}\n\nDocument text:\n{text}"
    try:
        raw = llm.classify_batch_json(full_prompt)
    except ValueError as exc:
        raise GuidelinesExtractionError(f"LLM returned malformed JSON: {exc}") from exc

    sanitised = _sanitise_extraction(raw)
    try:
        return FunderGuidelinesExtraction.model_validate(sanitised)
    except ValidationError as exc:
        raise GuidelinesExtractionError(f"FunderGuidelinesExtraction validation: {exc}") from exc


def _extract_from_image(
    image_bytes: bytes,
    media_subtype: str,
    llm: Any,  # noqa: ANN401 — BedrockClient duck-typed for test stubs
) -> Any:  # noqa: ANN401 — returns FunderGuidelinesExtraction
    """Extract guidelines from a single image via the vision path.

    Wraps ``BedrockClient.extract_raw_json_from_images`` and reuses the
    canonical sanitiser. ``media_subtype`` is informational — Bedrock
    accepts ``image/png``, ``image/jpeg``; PNG is the safe default for
    the streaming envelope (matches parser/vision usage). We re-encode
    JPEG as JPEG to avoid PIL roundtrips.
    """
    from pydantic import ValidationError

    from aegis.funders.guidelines_extract import (
        GUIDELINES_EXTRACTION_PROMPT,
        FunderGuidelinesExtraction,
        GuidelinesExtractionError,
        _sanitise_extraction,
    )

    # Vision envelope guesses media type from bytes; ``media_subtype`` is
    # a placeholder for future media-type-aware routing.
    del media_subtype

    try:
        raw, truncated = llm.extract_raw_json_from_images(
            [image_bytes], GUIDELINES_EXTRACTION_PROMPT
        )
    except ValueError as exc:
        raise GuidelinesExtractionError(f"LLM returned malformed JSON: {exc}") from exc

    if truncated:
        raise GuidelinesExtractionError(
            "LLM extraction was truncated at max_tokens — funder guideline "
            "image exceeds the model's output budget."
        )

    sanitised = _sanitise_extraction(raw)
    try:
        return FunderGuidelinesExtraction.model_validate(sanitised)
    except ValidationError as exc:
        raise GuidelinesExtractionError(f"FunderGuidelinesExtraction validation: {exc}") from exc


def _extract_for_file(
    file_path: Path,
    llm: Any,  # noqa: ANN401 — BedrockClient duck-typed
) -> Any:  # noqa: ANN401 — FunderGuidelinesExtraction
    """Dispatch to the right Bedrock entry based on the file extension.

    Raises ``GuidelinesExtractionError`` on extraction failure (caller
    catches and routes to the per-funder error bucket).
    """
    from aegis.funders.guidelines_extract import (
        GuidelinesExtractionError,
        extract_guidelines_from_pdf,
    )

    ext = file_path.suffix.lower()
    file_bytes = file_path.read_bytes()

    if ext in _PDF_EXTS:
        return extract_guidelines_from_pdf(file_bytes, llm)

    if ext in _IMAGE_EXTS:
        media_subtype = "jpeg" if ext in {".jpg", ".jpeg"} else "png"
        return _extract_from_image(file_bytes, media_subtype, llm)

    if ext in _DOCX_EXTS:
        if ext == ".doc":
            raise GuidelinesExtractionError(
                "legacy .doc binary format not supported — convert to .docx or .pdf"
            )
        text = _docx_to_text(file_path)
        if not text.strip():
            raise GuidelinesExtractionError("docx contained no extractable text")
        return _extract_from_text(text, llm)

    if ext in _XLSX_EXTS:
        if ext == ".xls":
            raise GuidelinesExtractionError(
                "legacy .xls binary format not supported — convert to .xlsx or .pdf"
            )
        text = _xlsx_to_text(file_path)
        if not text.strip():
            raise GuidelinesExtractionError("xlsx contained no extractable text")
        return _extract_from_text(text, llm)

    if ext in _TXT_EXTS:
        try:
            text = file_path.read_text(encoding="utf-8")
        except UnicodeDecodeError as exc:
            raise GuidelinesExtractionError(f"txt file not utf-8: {exc}") from exc
        if not text.strip():
            raise GuidelinesExtractionError("txt file was empty")
        return _extract_from_text(text, llm)

    raise GuidelinesExtractionError(f"unsupported file extension: {ext}")


def _merge_extractions(
    extractions: list[Any],
) -> Any | None:  # noqa: ANN401 — FunderGuidelinesExtraction | None
    """Merge per-file extractions into one ``FunderGuidelinesExtraction``.

    Merge semantics:

    * **Scalar fields** (``min_revenue, min_fico, min_tib_months,
      max_positions, stacking_policy, max_advance_amount``) — first
      non-None wins. ``model_dump(exclude_none=True)`` skips None
      entries; presence in the dict means non-None.
    * **List fields** (``excluded_industries, excluded_states``) —
      dedup-union across all extractions, preserving each item's first
      occurrence (case-sensitive on the surface, mirroring the
      operator-facing presentation).
    * **``notes``** — first non-empty string wins. The Pydantic default
      is the empty string (always present in ``model_dump`` output),
      so empty is treated as "absent" rather than a real value.

    Returns ``None`` when the input list is empty OR the merged dict
    fails ``FunderGuidelinesExtraction.model_validate`` (defensive —
    each input was already validated when produced, but a bug in the
    merge could in theory produce an out-of-range field).
    """
    from aegis.funders.guidelines_extract import FunderGuidelinesExtraction

    if not extractions:
        return None

    merged: dict[str, Any] = {}

    for extraction in extractions:
        data = extraction.model_dump(exclude_none=True)
        for field, value in data.items():
            if field in _MERGE_LIST_FIELDS:
                existing = merged.get(field, [])
                existing_keys = {str(x) for x in existing}
                new_items = [x for x in value if str(x) not in existing_keys]
                if new_items:
                    merged[field] = existing + new_items
                elif field not in merged:
                    # Preserve the empty list so the model_dump round
                    # trip retains the field shape.
                    merged[field] = existing
            elif field == "notes":
                if not merged.get(field) and value:
                    merged[field] = value
            elif field not in merged:
                merged[field] = value

    try:
        return FunderGuidelinesExtraction.model_validate(merged)
    except Exception as exc:
        print(
            f"  ! merge_extractions: validation failed on merged payload: {exc}",
            file=sys.stderr,
        )
        return None


def _stacking_to_bool(stacking_policy: str | None) -> bool | None:
    """Map the FunderGuidelinesExtraction stacking_policy Literal to the
    live FunderRow.accepts_stacking bool. ``case_by_case`` is left as
    ``None`` so the operator's review UI shows the field as "needs
    operator input" rather than defaulting to allow or deny.
    """
    if stacking_policy == "allowed":
        return True
    if stacking_policy == "not_allowed":
        return False
    return None


def _build_live_column_writes(extraction: Any) -> dict[str, Any]:  # noqa: ANN401 — FunderGuidelinesExtraction
    """Map the staging extraction to the live FunderRow columns.

    Per the operator-approved exception in the module docstring, the
    script writes these columns directly on INSERT / UPDATE. Fields the
    extractor dropped (confidence < 0.5) stay missing — the caller
    folds this dict into the row payload only for keys that ARE present.
    """
    out: dict[str, Any] = {}
    if extraction.min_revenue is not None:
        # Live column ``min_monthly_revenue`` is the deployed name; the
        # extraction model carries it under ``min_revenue`` because that
        # mirrors the operator-facing extraction prompt language.
        out["min_monthly_revenue"] = extraction.min_revenue
    if extraction.min_fico is not None:
        out["min_credit_score"] = extraction.min_fico
    if extraction.min_tib_months is not None:
        out["min_months_in_business"] = extraction.min_tib_months
    if extraction.max_positions is not None:
        out["max_positions"] = extraction.max_positions

    accepts_stacking = _stacking_to_bool(extraction.stacking_policy)
    if accepts_stacking is not None:
        out["accepts_stacking"] = accepts_stacking

    # excluded_* arrive as list[str]; the live FunderRow columns are
    # ARRAY<text> in Postgres — supabase-py round-trips a Python list
    # cleanly. Empty lists are NOT written (the operator's prior values
    # should not be wiped by a sparse extraction).
    if extraction.excluded_states:
        out["excluded_states"] = list(extraction.excluded_states)
    if extraction.excluded_industries:
        out["excluded_industries"] = list(extraction.excluded_industries)

    return out


def _normalise_subfolder_name(name: str) -> str:
    """Lower-case + strip — matches how the in-prod ``funders.name``
    lookup is compared (case-insensitive in the matcher; trailing
    whitespace would create false-misses against existing rows).
    """
    return name.strip().lower()


# Folder-name → canonical funder-name aliases (2026-06-30).
#
# Filip's local OneDrive folders use casual short names (LAG, TMR, UCS,
# VCG, "Highland capital", "Swiftsource") while the AEGIS catalog
# carries the canonical names (Logic Advance, TMRNOW, United Capital
# Source, etc.). Without this map a re-sync creates duplicate funder
# rows because the case-insensitive name match misses.
#
# Add new entries here whenever a SCP'd folder name diverges from the
# canonical ``funders.name`` row. Keys are lower-case + stripped (same
# shape as ``_normalise_subfolder_name``); values are the canonical
# name verbatim as it should appear in the ``funders`` table.
_FOLDER_NAME_ALIASES: dict[str, str] = {
    "highland capital": "Highland Hill Capital",
    "swiftsource": "SwiftSource Funding",
    "tmr": "TMRNOW",
    "ucs": "United Capital Source",
    "vcg": "Velocity Capital Group",
    "lag": "Logic Advance",
}


def _resolve_funder_name(folder_name: str) -> str:
    """Map a local folder name to its canonical funder name when an
    alias exists; otherwise return the original (stripped) name.

    Applied at the top of the per-subfolder loop so the lookup AND the
    new-row insert both see the canonical name. Prevents the duplicate-
    row pattern from the 2026-06-30 sync.
    """
    key = folder_name.strip().lower()
    return _FOLDER_NAME_ALIASES.get(key, folder_name.strip())


def _sync(
    *,
    folder_path: Path,
    apply_writes: bool,
) -> dict[str, list[str]]:
    """One-shot folder→AEGIS sync. Returns a bucket roll-up dict."""
    from aegis.db import get_supabase
    from aegis.funders.guidelines_extract import GuidelinesExtractionError
    from aegis.llm import BedrockClient

    if not folder_path.exists() or not folder_path.is_dir():
        print(
            f"CONFIG ERROR: funders folder not found or not a directory: {folder_path}",
            file=sys.stderr,
        )
        sys.exit(EXIT_CONFIG)

    print(f"# Reading funders from local folder: {folder_path}", file=sys.stderr)
    subfolders = sorted(
        (p for p in folder_path.iterdir() if p.is_dir()),
        key=lambda p: p.name.lower(),
    )
    print(f"# Found {len(subfolders)} funder subfolders\n", file=sys.stderr)

    try:
        sb = get_supabase()
    except Exception as exc:
        print(f"CONFIG ERROR: supabase init failed: {exc}", file=sys.stderr)
        sys.exit(EXIT_CONFIG)

    existing_resp = (
        sb.table("funders").select("id,name,guidelines_data,guidelines_uploaded_at").execute()
    )
    existing_rows = cast(list[dict[str, Any]], existing_resp.data or [])
    existing_by_name: dict[str, dict[str, Any]] = {
        _normalise_subfolder_name(row["name"]): row for row in existing_rows
    }

    llm = BedrockClient()

    results: dict[str, list[str]] = {
        "added": [],
        "updated": [],
        "skipped_up_to_date": [],
        "skipped_no_file": [],
        "errors": [],
    }

    for sub in subfolders:
        name = _resolve_funder_name(sub.name)
        files = _get_all_supported_files(sub)
        if not files:
            print(f"  - {name}: no supported files in folder", file=sys.stderr)
            results["skipped_no_file"].append(name)
            continue

        folder_hash = _folder_hash(sub)
        existing = existing_by_name.get(_normalise_subfolder_name(name))
        stored_hash = None
        if existing:
            stored_hash = (existing.get("guidelines_data") or {}).get("_file_hash")
        if existing and stored_hash == folder_hash:
            print(
                f"  = {name}: up to date (hash unchanged, {len(files)} files)",
                file=sys.stderr,
            )
            results["skipped_up_to_date"].append(name)
            continue

        action_verb = "UPDATE" if existing else "INSERT"

        # ── Dry-run: show per-file plan + would-merge summary ──────
        if not apply_writes:
            print(f"  [{action_verb} dry-run] {name}:", file=sys.stderr)
            eligible = 0
            skipped = 0
            for f in files:
                size_label = _format_file_size(f.stat().st_size)
                type_label = _content_label_for(f)
                if f.suffix.lower() in _IMAGE_EXTS and f.stat().st_size > _IMAGE_MAX_BYTES:
                    print(
                        f"    {f.name} ({size_label}, {type_label}) → SKIPPED: "
                        "exceeds 5MB Bedrock vision limit",
                        file=sys.stderr,
                    )
                    skipped += 1
                else:
                    print(
                        f"    {f.name} ({size_label}, {type_label}) → would extract",
                        file=sys.stderr,
                    )
                    eligible += 1
            print(
                f"    would merge {eligible} files, {skipped} skipped",
                file=sys.stderr,
            )
            results["updated" if existing else "added"].append(name)
            continue

        # ── Apply path: extract each file, collect, merge ──────────
        extractions: list[Any] = []
        files_log: list[dict[str, Any]] = []
        for f in files:
            size_label = _format_file_size(f.stat().st_size)
            type_label = _content_label_for(f)
            if f.suffix.lower() in _IMAGE_EXTS and f.stat().st_size > _IMAGE_MAX_BYTES:
                print(
                    f"    skip {f.name} ({size_label}, {type_label}): "
                    "exceeds 5MB Bedrock vision limit",
                    file=sys.stderr,
                )
                files_log.append(
                    {
                        "name": f.name,
                        "ext": f.suffix.lower(),
                        "action": "skipped",
                        "reason": "image_exceeds_5mb_bedrock_limit",
                    }
                )
                continue
            try:
                ext_result = _extract_for_file(f, llm)
            except GuidelinesExtractionError as exc:
                print(
                    f"    err {f.name}: {type(exc).__name__}: {exc}",
                    file=sys.stderr,
                )
                files_log.append(
                    {
                        "name": f.name,
                        "ext": f.suffix.lower(),
                        "action": "error",
                        "reason": f"{type(exc).__name__}: {exc}",
                    }
                )
                continue
            except Exception as exc:
                print(
                    f"    err {f.name}: unexpected {type(exc).__name__}: {exc}",
                    file=sys.stderr,
                )
                traceback.print_exc(file=sys.stderr)
                files_log.append(
                    {
                        "name": f.name,
                        "ext": f.suffix.lower(),
                        "action": "error",
                        "reason": f"unexpected_{type(exc).__name__}",
                    }
                )
                continue
            extractions.append(ext_result)
            files_log.append(
                {
                    "name": f.name,
                    "ext": f.suffix.lower(),
                    "action": "extracted",
                    "fields_populated": ext_result.fields_populated_count,
                }
            )

        if not extractions:
            print(f"  ! {name}: all files failed extraction", file=sys.stderr)
            results["errors"].append(f"{name}: all_files_failed")
            continue

        merged = _merge_extractions(extractions)
        if merged is None:
            print(f"  ! {name}: merged extraction validation failed", file=sys.stderr)
            results["errors"].append(f"{name}: merge_validation_failed")
            continue

        payload = merged.model_dump(mode="json")
        payload["_file_hash"] = folder_hash
        payload["_folder_name"] = name
        payload["_files"] = files_log

        write_row: dict[str, Any] = {
            "guidelines_data": payload,
            "guidelines_uploaded_at": datetime.now(UTC).isoformat(),
        }
        write_row.update(_build_live_column_writes(merged))

        try:
            if existing:
                sb.table("funders").update(cast(Any, write_row)).eq("id", existing["id"]).execute()
                results["updated"].append(name)
                print(
                    f"  ↑ {name}: updated "
                    f"(hash={folder_hash[:12]}…, "
                    f"{len(extractions)}/{len(files)} files merged)",
                    file=sys.stderr,
                )
            else:
                insert_row: dict[str, Any] = {
                    "name": name,
                    "operator_status": "active",
                    **write_row,
                }
                sb.table("funders").insert(cast(Any, insert_row)).execute()
                # Audit-log row for the auto-add — operator-visible
                # surface so a surprise insert is traceable to this run.
                sb.table("audit_log").insert(
                    cast(
                        Any,
                        {
                            "actor": "system:folder_sync",
                            "action": "funder.auto_added_from_folder",
                            "subject_type": "funder",
                            "details": {
                                "name": name,
                                "files": [entry["name"] for entry in files_log],
                                "folder": str(sub),
                            },
                        },
                    )
                ).execute()
                results["added"].append(name)
                print(
                    f"  + {name}: added ({len(extractions)}/{len(files)} files merged)",
                    file=sys.stderr,
                )
        except Exception as exc:
            print(f"  ! {name}: persist failed: {exc}", file=sys.stderr)
            results["errors"].append(f"{name}: persist_{type(exc).__name__}")

    return results


def main() -> int:
    _load_dotenv()
    args = _parse_args()

    # Late import — settings construction triggers the data-residency
    # boot guard, which we want to honor before walking the filesystem.
    from aegis.config import get_settings

    settings = get_settings()
    folder_str = args.folder or settings.funders_folder_path
    folder_path = Path(folder_str)

    apply_writes = bool(args.apply)
    mode = "APPLY" if apply_writes else "DRY-RUN"
    print(f"# mode={mode}", file=sys.stderr)

    results = _sync(folder_path=folder_path, apply_writes=apply_writes)
    print("", file=sys.stderr)
    print(f"# RESULT mode={mode}", file=sys.stderr)
    for bucket in (
        "added",
        "updated",
        "skipped_up_to_date",
        "skipped_no_file",
        "errors",
    ):
        print(f"  {bucket}: {len(results[bucket])}", file=sys.stderr)
    if results["errors"]:
        for err in results["errors"]:
            print(f"    - {err}", file=sys.stderr)
        return EXIT_PER_FUNDER_ERRORS
    return EXIT_OK


if __name__ == "__main__":
    sys.exit(main())
